from openpyxl import load_workbook

workbook = load_workbook('resource/file_example_XLSX_  50.xlsx')
sheet = workbook.active
print(sheet.cell(row=2, column=3).value)
print(f'Колличество листов{workbook.worksheets}')


